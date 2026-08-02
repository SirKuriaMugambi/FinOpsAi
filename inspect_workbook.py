from openpyxl import load_workbook

wb = load_workbook(
    'reference-data/CA- AI Payroll automation project.xlsx',
    read_only=True,
    data_only=True,
    password='8489'
)

print('SHEETS:', wb.sheetnames)
ws = wb['Integrating with AX cost center']

print('MAX_ROW:', ws.max_row)
print('MAX_COL:', ws.max_column)

for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
    print(row)
